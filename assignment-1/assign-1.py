import openpyxl
path=r"D:\myfiles\assign-1.xlsx"
wb=openpyxl.load_workbook(path)
ws=wb.active

for c1,c2,c3,c4 in ws['A1':'D7']:
    print("{0:7} {1:7} {2:7} {3:7}".format(c1.value,c2.value,c3.value, c4.value))

print()

print()

y=ws['D']

for i in range(1,len(y)):
    m=str(y[i].value)
    if m.isdigit():
        print(y[i].value/7)
    elif m.isalpha():
        print(y[i].value)  
    elif m.isdecimal():
        print(y[i].value)
    else:
        print()



